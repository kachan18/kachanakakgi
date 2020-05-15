import discord
import asyncio
import random
import datetime
import openpyxl
import os

client = discord.Client()


@client.event
async def on_ready():
    print("========================================")
    print("ID : ", client.user.id)
    print("Project Bot. Akagi")
    print("========================================")
    print("ready")
    game = discord.Game("지휘관 감시")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event
async def on_message(message):
    def asking(m):
        return m.channel == channel and m.author == m.author

    id = message.author.id
    channel = message.channel

    if message.author.bot:  # 봇은 기본적으로는 무시
        return None

    # and str(channel) == "아카기봇채널"
    if message.content.startswith("!아카기"):
        cmdline = message.content.split(' ')

        if len(cmdline) == 1:
            rand = random.randint(0, 100)
            if rand <= 45:
                await channel.send("네~지휘관님?")
            elif rand <= 95:
                await channel.send("우후후...이 아카기를 부르셨나요 지휘관님?")
            else:
                await channel.send("후후후...지휘관님을 유혹하는 눈엣가시들이 꽤나 많네요... 일단 「청소」를...\n어머, 아카기를 부르셨나요? 네? 별 일 아니랍니다...우후후")

        elif len(cmdline) >= 2:

            if cmdline[1] == "도움말":
                if len(cmdline) == 3 and cmdline[2] == "2":
                    embedhelp = discord.Embed(title="도움말을 찾으시나요, 지휘관님?", description="이 아카기가 전.부. 알려드리겠습니다.",
                                              color=0xf15f5f)
                    embedhelp.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    await channel.send(embed=embedhelp)
                else:
                    embedhelp = discord.Embed(title="도움말을 찾으시나요, 지휘관님?", description="이 아카기가 전.부. 알려드리겠습니다.",
                                              color=0xf15f5f)
                    embedhelp.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    embedhelp.add_field(name="!아카기", value="언제나 부르시길 기다리고 있습니다, 지휘관님♡", inline=False)
                    embedhelp.add_field(name="!아카기 정보", value="봇에 대한 정보를 열람합니다.", inline=False)
                    embedhelp.add_field(name="!아카기 도움말", value="도움말을 열람합니다.", inline=False)
                    embedhelp.add_field(name="!아카기 함선 (함선명)", value="함선에 대한 정보를 열람합니다.", inline=False)
                    embedhelp.add_field(name="!아카기 드랍 (함선명)", value="함선의 드랍 위치를 열람합니다. 해역에서만 드랍되는 함선만 사용 가능합니다.", inline=False)
                    embedhelp.add_field(name="!아카기 하드정보 지역명(예:3-4)", value="어려움 해역의 평균레벨 등 조건을 확인합니다.", inline=False)
                    embedhelp.add_field(name="!아카기 기억하기", value="지휘관님, 아카기에게 뭘 가르치고 싶으신가요?", inline=False)
                    embedhelp.add_field(name="!아카기 잊기", value="지휘관님, 아카기가 뭘 잊어버리기를 바라시나요?", inline=False)
                    embedhelp.add_field(name="!아카기 기억목록", value="지휘관님, 아카기가 뭘 기억하고 있는지 궁금하신가요?", inline=False)
                    embedhelp.add_field(name="!아카기 가위바위보", value="지휘관님, 아카기랑 가위바위보라도 해보시겠어요?", inline=False)
                    await channel.send(embed=embedhelp)

            elif cmdline[1] == "정보":
                embed = discord.Embed(title="아카기봇 v0.2",
                                      description="아카기봇\n24시간 가동 체제로 변경!\n아직 부족하지만 지휘관님과 함께라면 더욱 진화해 보이겠사와요.",
                                      color=0xf15f5f)
                embed.set_image(
                    url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                embed.add_field(name="```제작자```", value="Admiral. 레이나", inline=False)
                await channel.send(embed=embed)

            elif cmdline[1] == "드랍":
                file = openpyxl.load_workbook("드랍.xlsx")
                drop = file.active
                if len(cmdline) >= 3:
                    if random.randint(0, 100) <= 10 and cmdline[2] != "아카기":
                        await channel.send(
                            "```md\n[지휘관님?][지금 다른 아이에 대해 물어보셨죠? 아카기가 있으시면서, 어째서 다른 아이에 대해 물어보시는거죠? 유혹 당하셨나요? 그럼 그딴 녀석, 제가...]\n```")
                    else:
                        shipexist = 0
                        for i in range(2, 1000):
                            if drop["A" + str(i)].value == "-":
                                break
                            if cmdline[2] == drop["A" + str(i)].value:
                                shipexist = 1
                                embed = discord.Embed(title="%s 드랍" % drop["B" + str(i)].value,
                                                      description="%s" % drop["C" + str(i)].value,
                                                      color=0xf15f5f)
                                if drop["D" + str(i)].value != "-":
                                    embed.set_thumbnail(
                                        url="%s" % drop["D" + str(i)].value)
                                await channel.send(embed=embed)
                                break
                        if shipexist == 0:
                            await channel.send("그 함선에 대해서는 잘 모르는 것이와요.")
                else:
                    await channel.send("후후...지휘관님, 주신 정보가 부족하네요.")

            elif cmdline[1] == "함선":
                file = openpyxl.load_workbook("함선.xlsx")
                sheet = file.active

                if len(cmdline) >= 3:
                    if random.randint(0, 100) <= 10 and cmdline[2] != "아카기" and cmdline[2] != "카가" and cmdline[2] != "아마기":
                        await channel.send(
                            "```md\n[지휘관님?][지금 다른 아이에 대해 물어보셨죠? 아카기가 있으시면서, 어째서 다른 아이에 대해 물어보시는거죠? 유혹 당하셨나요? 그럼 그딴 녀석, 제가...]\n```")
                    else:
                        shipexist = 0
                        for i in range(2, 500):
                            if sheet["A" + str(i)].value == "-":
                                break
                            if cmdline[2] == sheet["O" + str(i)].value != "-" or cmdline[2] == sheet["P" + str(i)].value != "-":
                                shipexist = 1
                                embed = discord.Embed(title="%s" % sheet["B" + str(i)].value,
                                                      description="%s" % sheet["C" + str(i)].value,
                                                      color=0xf15f5f)
                                if sheet["D" + str(i)].value != "-":
                                    embed.set_image(
                                        url="%s" % sheet["D" + str(i)].value)
                                embed.add_field(name="```함종```", value="%s" % sheet["E" + str(i)].value, inline=False)
                                embed.add_field(name="```등급```", value="%s" % sheet["F" + str(i)].value, inline=False)
                                embed.add_field(name="```성우```", value="%s" % sheet["G" + str(i)].value, inline=False)
                                if sheet["L" + str(i)].value != "-":
                                    embed.add_field(name="```특이 사항```",
                                                    value="%s" % sheet["L" + str(i)].value,
                                                    inline=False)
                                if sheet["H" + str(i)].value != "-":
                                    embed.add_field(name="```스킬 1```",
                                                    value="%s" % sheet["H" + str(i)].value,
                                                    inline=False)
                                if sheet["I" + str(i)].value != "-":
                                    embed.add_field(name="```스킬 2```",
                                                    value="%s" % sheet["I" + str(i)].value,
                                                    inline=False)
                                if sheet["J" + str(i)].value != "-":
                                    embed.add_field(name="```스킬 3```",
                                                    value="%s" % sheet["J" + str(i)].value,
                                                    inline=False)
                                if sheet["K" + str(i)].value != "-":
                                    embed.add_field(name="```스킬 4```",
                                                    value="%s" % sheet["K" + str(i)].value,
                                                    inline=False)
                                await channel.send(embed=embed)
                                break
                            elif cmdline[2] == sheet["A" + str(i)].value:
                                if sheet["M" + str(i)].value == "-" or (len(cmdline) > 3 and cmdline[3] == sheet["M" + str(i)].value):
                                    if sheet["N" + str(i)].value == "-" or (len(cmdline) > 4 and cmdline[4] == sheet["N" + str(i)].value):
                                        shipexist = 1
                                        embed = discord.Embed(title="%s" % sheet["B" + str(i)].value,
                                                              description="%s" % sheet["C" + str(i)].value,
                                                              color=0xf15f5f)
                                        if sheet["D" + str(i)].value != "-":
                                            embed.set_image(
                                                url="%s" % sheet["D" + str(i)].value)
                                        embed.add_field(name="```함종```", value="%s" % sheet["E" + str(i)].value,
                                                        inline=False)
                                        embed.add_field(name="```등급```", value="%s" % sheet["F" + str(i)].value,
                                                        inline=False)
                                        embed.add_field(name="```성우```", value="%s" % sheet["G" + str(i)].value,
                                                        inline=False)
                                        if sheet["L" + str(i)].value != "-":
                                            embed.add_field(name="```특이 사항```",
                                                            value="%s" % sheet["L" + str(i)].value,
                                                            inline=False)
                                        if sheet["H" + str(i)].value != "-":
                                            embed.add_field(name="```스킬 1```",
                                                            value="%s" % sheet["H" + str(i)].value,
                                                            inline=False)
                                        if sheet["I" + str(i)].value != "-":
                                            embed.add_field(name="```스킬 2```",
                                                            value="%s" % sheet["I" + str(i)].value,
                                                            inline=False)
                                        if sheet["J" + str(i)].value != "-":
                                            embed.add_field(name="```스킬 3```",
                                                            value="%s" % sheet["J" + str(i)].value,
                                                            inline=False)
                                        if sheet["K" + str(i)].value != "-":
                                            embed.add_field(name="```스킬 4```",
                                                            value="%s" % sheet["K" + str(i)].value,
                                                            inline=False)
                                        await channel.send(embed=embed)
                                        break
                        if shipexist == 0:
                            await channel.send("잘 모르는 함선인 것이와요.")
                else:
                    await channel.send("후후...지휘관님, 어떤 아이에 대해 물어보시려고 하셨던 것인가요? 말해주신다면 이 아카기 잠시 「청소」를 하러... ")

            elif cmdline[1] == "하드정보":
                file = openpyxl.load_workbook("하드정보.xlsx")
                hard = file.active
                hardexist = 0

                if len(cmdline) >= 3:
                    for i in range(2, 30):
                        if hard["A" + str(i)].value == "-":
                            break
                        if cmdline[2] == hard["A" + str(i)].value:
                            hardexist = 1
                            embed = discord.Embed(title="어려움 %s" % hard["A" + str(i)].value,
                                                  description="%s" % hard["B" + str(i)].value,
                                                  color=0xf15f5f)
                            embed.add_field(name="```1함대 조건```", value="%s" % hard["C" + str(i)].value, inline=False)
                            embed.add_field(name="```2함대 조건```", value="%s" % hard["D" + str(i)].value, inline=False)
                            if hard["E" + str(i)].value != "-":
                                if hard["F" + str(i)].value != "-":
                                    if hard["G" + str(i)].value != "-":
                                        embed.add_field(name="```출격 조건```", value="%s / %s / %s" % (hard["E" + str(i)].value, hard["F" + str(i)].value, hard["G" + str(i)].value), inline=False)
                            embed.add_field(name="```적 정보```",
                                            value="정보 없음",
                                            inline=False)
                            await channel.send(embed=embed)
                            break
                    if hardexist == 0:
                        await channel.send("아직 해당 지역의 정보가 없는 것이와요.")
                else:
                    await channel.send("후후...지휘관님, 주신 정보가 부족하네요.")

            elif cmdline[1] == "기억하기":
                file = openpyxl.load_workbook("기억리스트.xlsx")
                remember = file.active
                remnum = 101
                already = 0
                isfull = 0

                embed = discord.Embed(title="지휘관님 무엇을 말하실때 답할까요?",
                                      description="10초 이내로 답해주세요. 그냥 채팅으로 치시면 됩니다.",
                                      color=0xf15f5f)
                embed.set_thumbnail(
                    url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                await channel.send(embed=embed)
                try:
                    msg = await client.wait_for('message', timeout=10.0, check=asking)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="\n10초안에 알려주시기로 하셨는데 늦으셨네요.",
                                          description="다른 생각이라도 하시고 계셨나요 지휘관님?\n혹시 다른 아이 생각이라도......저기 지휘관님, 누구야? 누굴 생각한거야?",
                                          color=0xf15f5f)
                    embed.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    await channel.send(embed=embed)
                else:
                    for i in range(1, 101):
                        isfull = 1
                        if remember["A" + str(i)].value == "-":
                            remember["A" + str(i)] = msg.content
                            isfull = 0
                            remnum = i
                            break
                    if isfull == 0:
                        for j in range(1, remnum):
                            already = 0
                            if remember["A" + str(j)].value == remember["A" + str(remnum)].value:
                                already = 1
                                break
                        if already == 1:
                            embed = discord.Embed(title="지휘관님이 요청하신 내용은 이미 있는 질문이랍니다.",
                                                  description="안타깝게도 요청이 취소되었어요",
                                                  color=0xf15f5f)
                            embed.set_thumbnail(
                                url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                            await channel.send(embed=embed)
                            remember["A" + str(remnum)] = "-"
                            file.save("기억리스트.xlsx")
                        else:
                            embed = discord.Embed(title="\"%s\"라고 하시면 \n뭐라고 답할까요?" % msg.content,
                                                  description="10초 이내로 답해주세요. 그냥 채팅으로 치시면 됩니다.",
                                                  color=0xf15f5f)
                            embed.set_thumbnail(
                                url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                            await channel.send(embed=embed)
                            try:
                                msg = await client.wait_for('message', timeout=10.0, check=asking)
                            except asyncio.TimeoutError:
                                embed = discord.Embed(title="\n10초안에 알려주시기로 하셨는데 늦으셨네요.",
                                                      description="다른 생각이라도 하시고 계셨나요 지휘관님?\n혹시 다른 아이 생각이라도......저기 지휘관님, 누구야? 누굴 생각한거야?",
                                                      color=0xf15f5f)
                                embed.set_thumbnail(
                                    url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                                await channel.send(embed=embed)
                                remember["A" + str(remnum)] = "-"
                                file.save("기억리스트.xlsx")
                            else:
                                remember["B" + str(remnum)] = msg.content
                                file.save("기억리스트.xlsx")
                                embed = discord.Embed(title="지휘관님의 말 잘 기억하고 있겠어요.",
                                                      description="!아카기 대답 (질문) 으로 물어보실 수 있습니다.",
                                                      color=0xf15f5f)
                                embed.set_thumbnail(
                                    url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                                embed.add_field(name="```질문```", value="%s" % remember["A" + str(remnum)].value,
                                                inline=False)
                                embed.add_field(name="```대답```", value="%s" % remember["B" + str(remnum)].value,
                                                inline=False)
                                await channel.send(embed=embed)
                    else:
                        embed = discord.Embed(title="이미 사용 가능한 질문의 수가 꽉 차버렸답니다.",
                                              description="안타깝게도 요청이 취소되었어요",
                                              color=0xf15f5f)
                        embed.set_thumbnail(
                            url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                        await channel.send(embed=embed)
            elif cmdline[1] == "대답":
                file = openpyxl.load_workbook("기억리스트.xlsx")
                remember = file.active
                findanswer = 0
                ask = []

                for i in range(2, len(cmdline)):
                    ask.append(cmdline[i])
                ask = " ".join(ask)

                for i in range(1, 101):
                    if remember["A" + str(i)].value == ask:
                        findanswer = 1
                        embed = discord.Embed(title="%s" % remember["B" + str(i)].value,
                                              description="질문 : \"%s\"" % ask,
                                              color=0xf15f5f)
                        embed.set_thumbnail(
                            url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                        await channel.send(embed=embed)
                        break
                if findanswer == 0:
                    embed = discord.Embed(title="뭐라고 답해야 할지 모르겠네요",
                                          description="질문 : \"%s\"" % ask,
                                          color=0xf15f5f)
                    embed.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    await channel.send(embed=embed)
            elif cmdline[1] == "기억목록":
                file = openpyxl.load_workbook("기억리스트.xlsx")
                remember = file.active
                isempty = 1
                embed = discord.Embed(title="아카기의 기억 목록",
                                      description="제 기억이 궁금하셨나요 지휘관님? 지휘관님을 위해서라면.",
                                      color=0xf15f5f)
                for i in range(1, 100):
                    if remember["A" + str(i)].value != "-":
                        embed.add_field(name="```기억 %d```" % i, value="질문 : \"%s\"\n대답 : \"%s\"" % (remember["A" + str(i)].value, remember["B" + str(i)].value), inline=False)
                        isempty = 0
                if isempty == 1:
                    embed = discord.Embed(title="아카기의 기억 목록",
                                          description="현재 기억하고 있는 내용이 없네요",
                                          color=0xf15f5f)
                    embed.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    await channel.send(embed=embed)
                else:
                    embed.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    await channel.send(embed=embed)

            elif cmdline[1] == "잊기":
                file = openpyxl.load_workbook("기억리스트.xlsx")
                remember = file.active
                isempty = 1
                embed = discord.Embed(title="아카기의 기억 목록",
                                      description="어떤 기억을 잊기를 원하시나요 지휘관님? 번호로 답해주세요.",
                                      color=0xf15f5f)
                for i in range(1, 100):
                    if remember["A" + str(i)].value != "-":
                        embed.add_field(name="```기억 %d```" % i, value="질문 : \"%s\"\n대답 : \"%s\"" % (remember["A" + str(i)].value, remember["B" + str(i)].value), inline=False)
                        isempty = 0
                if isempty == 1:
                    embed = discord.Embed(title="아카기의 기억 목록",
                                          description="현재 기억하고 있는 내용이 없네요",
                                          color=0xf15f5f)
                    embed.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    await channel.send(embed=embed)
                else:
                    embed.set_thumbnail(
                        url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                    await channel.send(embed=embed)

                    def asking(m):
                        return m.channel == channel and m.author == m.author

                    try:
                        msg = await client.wait_for('message', timeout=5.0, check=asking)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="\n지휘관님이 생각을 바꾸셨나봐요.",
                                              description="혹시 잊어버리셨나요, 지휘관님?",
                                              color=0xf15f5f)
                        embed.set_thumbnail(
                            url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                        await channel.send(embed=embed)
                    else:
                        if remember["A" + str(msg.content)].value != "-":
                            embed = discord.Embed(title="\n지휘관님이 원하시는대로 잊어버렸답니다",
                                                  description="잊은 내용```\n질문 : \"%s\"\n대답 : \"%s\"```" % (remember["A" + str(msg.content)].value, remember["B" + str(msg.content)].value),
                                                  color=0xf15f5f)
                            embed.set_thumbnail(
                                url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                            remember["A" + str(msg.content)] = "-"
                            remember["B" + str(msg.content)] = "-"
                            file.save("기억리스트.xlsx")
                            await channel.send(embed=embed)
                        else:
                            embed = discord.Embed(title="\n없는 기억이네요",
                                                  description="생각을 바꾸셨나요, 지휘관님?",
                                                  color=0xf15f5f)
                            embed.set_thumbnail(
                                url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                            await channel.send(embed=embed)

            elif cmdline[1] == "가위바위보":
                file = openpyxl.load_workbook("가위바위보쿨다운.xlsx")
                sheet = file.active
                for i in range(1, 51):
                    if float(message.author.id) == sheet["A" + str(i)].value:
                        if int(sheet["B" + str(i)].value) <= int(datetime.datetime.now().strftime("%Y%m%d%H%M%S")):
                            cooldown = datetime.datetime.now() + datetime.timedelta(seconds=30)
                            sheet["B" + str(i)].value = cooldown.strftime("%Y%m%d%H%M%S")
                            file.save("가위바위보쿨다운.xlsx")
                            embed = discord.Embed(title="지휘관님 아카기랑 가위바위보 하시겠어요?\n10초안에 내시기에요?",
                                                  description="가위 바위 보 중 하나로 답해주세요",
                                                  color=0xf15f5f)
                            embed.set_thumbnail(
                                url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                            await channel.send(embed=embed)

                            def rcp(m):
                                return m.channel == channel and (m.content == '가위' or m.content == '바위' or m.content == '보') and m.author == m.author

                            try:
                                msg = await client.wait_for('message', timeout=10.0, check=rcp)
                            except asyncio.TimeoutError:
                                embed = discord.Embed(title="\n10초안에 내시기로 하셨는데 늦으셨네요",
                                                      description="다른 생각이라도 하시고 계셨나요 지휘관님?\n혹시 다른 아이 생각이라도......저기 지휘관님, 누구야? 누굴 생각한거야?",
                                                      color=0xf15f5f)
                                embed.set_thumbnail(
                                    url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                                await channel.send(embed=embed)
                            else:
                                rand = random.randint(0, 2)
                                rcpname = "가위 바위 보".split(' ')
                                await channel.send("```[ %s ]지휘관님은 [ %s ] 를 내셨네요.\n아카기는 무엇이냐면요,\n\n**%s 랍니다!**```" % (msg.author, msg.content, rcpname[rand]))
                                if msg.content == "가위":
                                    if rand == 0:
                                        await channel.send("```어머, 비겼네요. 이렇게 마음이 맞는걸 보면 역시, 아카기와 지휘관님은 맺어질 운명이에요...```")
                                    elif rand == 1:
                                        await channel.send("```후후, 아카기가 이겼네요. 아카기가 이겼으니까 지휘관님의 모든 것은 제 것이에요~♥```")
                                    else:
                                        await channel.send("```지휘관님이 이기셨으니 상품을 드릴게요. 상품은 바로, 아.카.기 랍니다? 이 아카기의 모든 것 소중히 여겨주세요```")
                                elif msg.content == '바위':
                                    if rand == 0:
                                        await channel.send("```지휘관님이 이기셨으니 상품을 드릴게요. 상품은 바로, 아.카.기 랍니다? 이 아카기의 모든 것 소중히 여겨주세요```")
                                    elif rand == 1:
                                        await channel.send("```어머, 비겼네요. 이렇게 마음이 맞는걸 보면 역시, 아카기와 지휘관님은 맺어질 운명이에요...```")
                                    else:
                                        await channel.send("```후후, 아카기가 이겼네요. 아카기가 이겼으니까 지휘관님의 모든 것은 제 것이에요~♥```")
                                else:
                                    if rand == 0:
                                        await channel.send("```후후, 아카기가 이겼네요. 아카기가 이겼으니까 지휘관님의 모든 것은 제 것이에요~♥```")
                                    elif rand == 1:
                                        await channel.send("```지휘관님이 이기셨으니 상품을 드릴게요. 상품은 바로, 아.카.기 랍니다? 이 아카기의 모든 것 소중히 여겨주세요```")
                                    else:
                                        await channel.send("```어머, 비겼네요. 이렇게 마음이 맞는걸 보면 역시, 아카기와 지휘관님은 맺어질 운명이에요...```")
                            break
                        else:
                            await channel.send("```[%s] 지휘관님, 아직 재사용 대기시간이 %d초 남았어요.```" % (message.author, (datetime.datetime.strptime(sheet["B" + str(i)].value, '%Y%m%d%H%M%S') - datetime.datetime.now()).seconds))
                            break
                    if sheet["A" + str(i)].value == "-":
                        sheet["A" + str(i)] = message.author.id
                        cooldown = datetime.datetime.now() + datetime.timedelta(seconds=30)
                        sheet["B" + str(i)] = cooldown.strftime("%Y%m%d%H%M%S")
                        file.save("가위바위보쿨다운.xlsx")
                        embed = discord.Embed(title="지휘관님 아카기랑 가위바위보 하시겠어요?\n10초안에 내시기에요?",
                                              description="가위 바위 보 중 하나로 답해주세요",
                                              color=0xf15f5f)
                        embed.set_thumbnail(
                            url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                        await channel.send(embed=embed)

                        def rcp(m):
                            return m.channel == channel and (
                                    m.content == '가위' or m.content == '바위' or m.content == '보') and m.author == m.author

                        try:
                            msg = await client.wait_for('message', timeout=10.0, check=rcp)
                        except asyncio.TimeoutError:
                            embed = discord.Embed(title="\n10초안에 내시기로 하셨는데 늦으셨네요",
                                                  description="다른 생각이라도 하시고 계셨나요 지휘관님?\n혹시 다른 아이 생각이라도......저기 지휘관님, 누구야? 누굴 생각한거야?",
                                                  color=0xf15f5f)
                            embed.set_thumbnail(
                                url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                            await channel.send(embed=embed)
                        else:
                            rand = random.randint(0, 2)
                            rcpname = "가위 바위 보".split(' ')
                            await channel.send("```[ %s ]지휘관님은 [ %s ] 를 내셨네요.\n아카기는 무엇이냐면요,\n\n**%s 랍니다!**```" % (msg.author, msg.content, rcpname[rand]))
                            if msg.content == "가위":
                                if rand == 0:
                                    await channel.send("```어머, 비겼네요. 이렇게 마음이 맞는걸 보면 역시, 아카기와 지휘관님은 맺어질 운명이에요...```")
                                elif rand == 1:
                                    await channel.send("```후후, 아카기가 이겼네요. 아카기가 이겼으니까 지휘관님의 모든 것은 제 것이에요~♥```")
                                else:
                                    await channel.send(
                                        "```지휘관님이 이기셨으니 상품을 드릴게요. 상품은 바로, 아.카.기 랍니다? 이 아카기의 모든 것 소중히 여겨주세요```")
                            elif msg.content == '바위':
                                if rand == 0:
                                    await channel.send(
                                        "```지휘관님이 이기셨으니 상품을 드릴게요. 상품은 바로, 아.카.기 랍니다? 이 아카기의 모든 것 소중히 여겨주세요```")
                                elif rand == 1:
                                    await channel.send("```어머, 비겼네요. 이렇게 마음이 맞는걸 보면 역시, 아카기와 지휘관님은 맺어질 운명이에요...```")
                                else:
                                    await channel.send("```후후, 아카기가 이겼네요. 아카기가 이겼으니까 지휘관님의 모든 것은 제 것이에요~♥```")
                            else:
                                if rand == 0:
                                    await channel.send("```후후, 아카기가 이겼네요. 아카기가 이겼으니까 지휘관님의 모든 것은 제 것이에요~♥```")
                                elif rand == 1:
                                    await channel.send(
                                        "```지휘관님이 이기셨으니 상품을 드릴게요. 상품은 바로, 아.카.기 랍니다? 이 아카기의 모든 것 소중히 여겨주세요```")
                                else:
                                    await channel.send("```어머, 비겼네요. 이렇게 마음이 맞는걸 보면 역시, 아카기와 지휘관님은 맺어질 운명이에요...```")
                        break

            elif cmdline[1] == "서약":
                embed = discord.Embed(title="아카기, 이날만을 기다렸어요.\n후후훗...... 앞으로는 누구든 간에 지휘관님과 아카기를 떨어뜨릴 수 없어.\n저의 모든 것이 지휘관님의 것, 지휘관님의 모든 것은 저의 것이에요..... 후훗, 후후훗, 우후후후후후훗......",
                                      description="\n",
                                      color=0xf15f5f)
                embed.set_image(
                    url="https://mblogthumb-phinf.pstatic.net/MjAxODA1MjNfMTQy/MDAxNTI3MDYzOTUxMDgy.WjbntAVv-ZJ_M4-7l_F5yDWu56LpUMtZs-Fwfwcy5sUg.owTwDTBHq6KO-feCwYNLJNw_wPV7v1fvDWfoztzxiscg.JPEG.hcnhd/Screenshot_20180523-171244.jpg?type=w800")
                await channel.send(embed=embed)

            elif cmdline[1] == "랜덤대사":
                file = openpyxl.load_workbook("랜덤대사.xlsx")
                sheet = file.active

                embed = discord.Embed(title="후후, 아카기가 말하는게 듣고싶으셨나요?\n지휘관님이 원하신다면...",
                                      description="",
                                      color=0xf15f5f)
                embed.set_thumbnail(
                    url="https://images2.imgbox.com/a5/cd/nXI2XWKF_o.png")
                randlen = 1
                for i in range(1, 100):
                    if sheet["A" + str(i)].value != "-":
                        randlen += 1
                    else:
                        break
                rand = random.randint(1, randlen)
                embed.add_field(name="%s" % sheet["A" + str(rand)].value, value="%s" % sheet["B" + str(rand)].value, inline=False)
                await channel.send(embed=embed)
            else:
                await channel.send("지휘관님이 무엇을 말하시려 했는지는 모르겠지만, 지휘관님은 아카기와 영원히 함께랍니다~♥")
    if message.content.startswith("!카가"):
        await message.channel.send("어머, 카가는 이곳에 없답니다, 지휘관님?")
    if message.content.startswith("!아마기"):
        await message.channel.send("어머, 언니께서는 이곳에 없답니다, 지휘관님?")

access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
